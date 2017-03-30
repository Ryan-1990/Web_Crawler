# -*- coding: utf-8 -*-
#Author: Nan LI @ Tianjin University
#功能描述：遍历寻医问药网上所有地区的医院，并按地区分类将医院名称及标签记录到一个文件里

import sys
import time
import urllib2
import requests
import lxml.html
import random
import string
import re
import xlrd
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


# function: get the directory of script that calls this function
# usage: path = script_path()
def script_path():
    import inspect, os
    caller_file = inspect.stack()[1][1]         # caller's filename
    return os.path.abspath(os.path.dirname(caller_file))# path

def create_folder():
    import os
    path = script_path()
    title = "Xun_Yi_Wen_Yao"
    new_path = os.path.join(path, title)
    if not os.path.isdir(new_path):
        os.makedirs(new_path)
        

class Search:
    def __init__(self, address):
        #print address
        self.address = address

        randnum = random.randint(1,9)
        if randnum == 1:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/33.0.1750.154 Safari/537.36'}
        elif randnum == 2:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)',
                           'Referer': 'http://www.baidu.com'}
        elif randnum == 3:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Win64; x64; Trident/4.0)'}
        elif randnum == 4:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)'}
        elif randnum == 5:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0)'}
        elif randnum == 6:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1'}
        elif randnum == 7:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12'}
        elif randnum == 8:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; .NET CLR 2.0.50727; Maxthon 2.0'}
        else:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) AppleWebKit/525.13 (KHTML, like Gecko) Version/3.1 Safari/525.13'}

        #sleepseconds = random.uniform(0, 2)   # Sleep randomly from 0 sec to 5.0 sec. before launching another crawling.
        #time.sleep(sleepseconds)

        req = requests.Request(method='GET',
                               url=self.address,
                               headers=self.header,
                               cookies=None)
        reqprep = req.prepare()
        s = requests.Session()
        resp = s.send(reqprep)

        self.text = resp.text
        self.tree = lxml.html.fromstring(resp.text)
        
class ExtractorWebsite(Search):
    def extractor_website(self):
        a = self.tree.xpath('//div[@class="bd f12 btn-a deepgray-a"]/div[@class="categories"]/a')
        if len(a) == 0:
            return False
        return a

    def extractor_city(self):
        a = self.tree.xpath('//div[@class="bd f12 btn-a deepgray-a"]/div[@class="categories"]/a//span')
        if len(a) == 0:
            return False
        return a

    def get_district(self):
        b = self.tree.xpath('//div[@class="caption sprite-repeat pl20"]//h3')
        if len(b) == 0:
            return False
        return b

    def hospital_names(self,index):
        c = self.tree.xpath('//div[@class="pl15 pr5 pt15 pb10 f14 deepgray-a province"][%d]/ul[@class="clearfix"]/li/a'%(index+1))
        if len(c) == 0:
            return False
        return c

    def hospital_tags(self,index):
        c = self.tree.xpath('//div[@class="pl15 pr5 pt15 pb10 f14 deepgray-a province"][%d]/ul[@class="clearfix"]/li/span'%(index+1))
        if len(c) == 0:
            return False
        return c

    def district_names(self,index):
        d = self.tree.xpath('//div[@class="caption sprite-repeat pl20"][%d]//h3'%(index+1))
        if len(d) == 0:
            return False
        return d

    def get_phone(self):
        a = self.tree.xpath('//head/meta[3]')
        if len(a) == 0:
            return False
        return a

    def get_intro(self):
        a = self.tree.xpath('//dl[@class="clearfix"]/dd[@class=" fl"]//a')
        if len(a) == 0:
            return False
        return a

    def get_department(self):
        a = self.tree.xpath('//ul[@class="clearfix f14 "]/li[@class="fl tc sprite-repeat brdc8 mr5"][1]//a')
        if len(a) == 0:
            return False
        return a

    def get_expert(self):
        a = self.tree.xpath('//ul[@class="clearfix f14 "]/li[@class="fl tc sprite-repeat brdc8 mr5"][3]//a')
        if len(a) == 0:
            return False
        return a


def Start():
    create_folder()
    ScriptPath = script_path()
    XYWY_Path = ScriptPath+"/Xun_Yi_Wen_Yao/"
    #---------新建一个文件--------
    wb1 = Workbook()
    ewb1 = ExcelWriter(workbook=wb1)
    new_filename=XYWY_Path+'/Hospital_Names.xlsx'
    ws1=wb1.worksheets[0]
    ws1.title="Sheet1"
    ws1.cell(row=0,column=0).value = "省市"
    ws1.cell(row=0,column=1).value = "城区"
    ws1.cell(row=0,column=2).value = "医院名称"
    ws1.cell(row=0,column=3).value = "医院标签"
    ws1.cell(row=0,column=4).value = "医院链接"
    ws1.cell(row=0,column=5).value = "医院地址"
    ws1.cell(row=0,column=6).value = "医院电话"
    ws1.cell(row=0,column=7).value = "介绍页面链接"
    ws1.cell(row=0,column=8).value = "科室链接"
    ws1.cell(row=0,column=9).value = "推荐专家链接"
    
    XYWY_address = "http://z.xywy.com/yiyuan.htm"
    city_name = []
    city_address = []
    district_cnt = []

    copy = ExtractorWebsite(XYWY_address)
    href_list = copy.extractor_website()
    city_list = copy.extractor_city()
    count = len(href_list)
    row_num = 0
    for cnt in xrange(0,count):
    #for cnt in xrange(0,1):
        city_name.append(city_list[cnt].text.encode('utf8'))
        print city_name[cnt]
        for href in href_list[cnt].items():
            if href[0] == 'href':
                city_address.append(href[1])
                print city_address[cnt]
        ########################################################################
        district_name = []
        copy = ExtractorWebsite(city_address[cnt])
        temp = copy.get_district()
        district_cnt.append(len(temp))
        print district_cnt[cnt]
        for dis_cnt in xrange(0,district_cnt[cnt]):
            dis_names = copy.district_names(dis_cnt)
            district_name.append(dis_names[0].text)
            print district_name[dis_cnt]
            
            names = copy.hospital_names(dis_cnt)
            tags = copy.hospital_tags(dis_cnt)
            if names:
                name_cnt = len(names)
                for namecnt in range(0,name_cnt):
                    name = names[namecnt].text
                    tag = tags[namecnt].text
                    print name+tag
                    
                    for href in names[namecnt].items():
                        if href[0] == 'href':
                            hospital_address = href[1]
                            print hospital_address
                               
                    row_num+=1
                    ws1.cell(row=row_num,column=0).value = city_name[cnt]
                    ws1.cell(row=row_num,column=1).value = district_name[dis_cnt]
                    ws1.cell(row=row_num,column=2).value = name
                    ws1.cell(row=row_num,column=3).value = tag+u"）"
                    ws1.cell(row=row_num,column=4).value = hospital_address

                    copy1 = ExtractorWebsite(hospital_address)
                    temp = copy1.get_phone()
                    for content in temp[0].items():
                        if content[0] == 'content':
                            content_str = content[1].encode('utf8')
                    #print content_str
                    Address = re.search('地址.*;',content_str)
                    print Address.group()
                    Phone = re.search('电话.*',content_str)
                    print Phone.group()

                    temp = copy1.get_intro()
                    for href in temp[0].items():
                        if href[0] == 'href':
                            intro_address = href[1]
                            print intro_address

                    temp = copy1.get_department()
                    for href in temp[0].items():
                        if href[0] == 'href':
                            department_address = href[1]
                            print department_address

                    temp = copy1.get_expert()
                    for href in temp[0].items():
                        if href[0] == 'href':
                            expert_address = href[1]
                            print expert_address

                    ws1.cell(row=row_num,column=5).value = Address.group()
                    ws1.cell(row=row_num,column=6).value = Phone.group()
                    ws1.cell(row=row_num,column=7).value = intro_address
                    ws1.cell(row=row_num,column=8).value = department_address
                    ws1.cell(row=row_num,column=9).value = expert_address
                    #ewb1.save(filename=new_filename)
                    
            else:
                print "没有找到医院"
                #print 0
                row_num+=1
                ws1.cell(row=row_num,column=0).value = city_name[cnt]
                ws1.cell(row=row_num,column=1).value = "没有找到医院"
                ws1.cell(row=row_num,column=2).value = "没有找到医院"
                ws1.cell(row=row_num,column=3).value = "None"
                ws1.cell(row=row_num,column=4).value = "None"
                
            ewb1.save(filename=new_filename)
        

if __name__ == "__main__":
    Start()
