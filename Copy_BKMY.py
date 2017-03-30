# -*- coding: utf-8 -*-
#Author: Nan LI @ Tianjin University
#功能描述：遍历百科名医网的所有科室，并遍历每个科室的所有疾病，最后把所有疾病名称记录到一个文件里

import sys
import time
import urllib2
import requests
import lxml.html
import random
import string
import re
import xlrd
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


# function: get the directory of script that calls this function
# usage: path = script_path()
def script_path():
    import inspect, os
    caller_file = inspect.stack()[1][1]         # caller's filename
    return os.path.abspath(os.path.dirname(caller_file))# path

def create_folder():
    import os
    path = script_path()
    title = "Bai_Ke_Ming_Yi"
    new_path = os.path.join(path, title)
    if not os.path.isdir(new_path):
        os.makedirs(new_path)
        

class Search:
    def __init__(self, address):
        #print address
        self.address = address

        randnum = random.randint(1,9)
        if randnum == 1:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/33.0.1750.154 Safari/537.36'}
        elif randnum == 2:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)',
                           'Referer': 'http://www.baidu.com'}
        elif randnum == 3:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Win64; x64; Trident/4.0)'}
        elif randnum == 4:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)'}
        elif randnum == 5:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0)'}
        elif randnum == 6:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1'}
        elif randnum == 7:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12'}
        elif randnum == 8:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; .NET CLR 2.0.50727; Maxthon 2.0'}
        else:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) AppleWebKit/525.13 (KHTML, like Gecko) Version/3.1 Safari/525.13'}

        #sleepseconds = random.uniform(0, 2)   # Sleep randomly from 0 sec to 5.0 sec. before launching another crawling.
        #time.sleep(sleepseconds)

        req = requests.Request(method='GET',
                               url=self.address,
                               headers=self.header,
                               cookies=None)
        reqprep = req.prepare()
        s = requests.Session()
        resp = s.send(reqprep)

        self.text = resp.text
        self.tree = lxml.html.fromstring(resp.text)
        
class ExtractorWebsite(Search):
    def get_url(self, items):
        for i in items:
           if i[0] == 'href':
              return i[1]
    
    def extractor_website(self):
        a = self.tree.xpath('//ul[@class="keshi_list clearFix"]/li/a')
        if len(a) == 0:
            return False
        return a

    def get_diseases(self):
        b = self.tree.xpath('//div[@class="page"]/a')
        if len(b) == 0:
            return False
        return b

    def disease_names(self):
        c = self.tree.xpath('//div[@class="jibing-list"]/ul[@class="clearFix"]/li/a')
        if len(c) == 0:
            return False
        return c

    def extractor_baike(self):
        d = self.tree.xpath('//div[@class="mod-list"]//a')
        if len(d) == 0:
            return False
        return self.get_url(d[0].items())

def Start():
    create_folder()
    ScriptPath = script_path()
    BKMY_Path = ScriptPath+"/Bai_Ke_Ming_Yi/"
    #---------新建一个文件--------
    wb1 = Workbook()
    ewb1 = ExcelWriter(workbook=wb1)
    new_filename=BKMY_Path+'/Disease_Names.xlsx'
    ws1=wb1.worksheets[0]
    ws1.title="Sheet1"
    ws1.cell(row=0,column=0).value = "所属科类别"
    ws1.cell(row=0,column=1).value = "疾病名称"
    ws1.cell(row=0,column=2).value = "百科名医链接"
    ws1.cell(row=0,column=3).value = "百度百科链接"
    ws1.cell(row=0,column=4).value = "全球医院链接"
    
    BKMY_address = "http://www.baikemy.com"
    department_name = []
    department_address = []
    page_cnt = []

    copy = ExtractorWebsite(BKMY_address+"/disease/list/0/0")
    href_list = copy.extractor_website()
    count = len(href_list)
    row_num = 0
    for cnt in xrange(0,count):
    #for cnt in xrange(0,1):
        department_name.append(href_list[cnt].text.encode('utf8'))
        print department_name[cnt]
        for href in href_list[cnt].items():
            if href[0] == 'href':
                department_address.append(BKMY_address+href[1])
                print department_address[cnt]
        ########################################################################
        copy = ExtractorWebsite(department_address[cnt])
        temp = copy.get_diseases() 
        if temp:
            page_cnt.append(len(temp)-1)
        else:
            page_cnt.append(1)
        print page_cnt[cnt]
        ########################################################################
        names = copy.disease_names()
        if names:
            name_cnt = len(names)
            for namecnt in range(0,name_cnt):
                name = names[namecnt].text
                print name
                
                for href in names[namecnt].items():
                    if href[0] == 'href':
                        disease_address = BKMY_address+href[1]
                        print disease_address
                        
                bk_address = "http://baike.baidu.com/search/word?word="
                bk_address += name
                copy11 = ExtractorWebsite(bk_address)
                baike_address = copy11.extractor_baike()
                print baike_address
                ##---------------------------------------------------##
                qqyy_address = "http://so.qqyy.com/j?wd=" 
                qqyy_address += name
                print qqyy_address
                ##---------------------------------------------------##    
       
                row_num+=1
                ws1.cell(row=row_num,column=0).value = department_name[cnt]
                ws1.cell(row=row_num,column=1).value = name
                ws1.cell(row=row_num,column=2).value = disease_address
                ws1.cell(row=row_num,column=3).value = baike_address
                ws1.cell(row=row_num,column=4).value = qqyy_address

            for page_num in xrange(2,page_cnt[cnt]+1):
                copy1 = ExtractorWebsite(department_address[cnt]+"?pageIndex=%d"%page_num)
                names = copy1.disease_names()
                name_cnt += len(names)
                for namecnt in range(0,len(names)):
                    name = names[namecnt].text
                    print name
                    for href in names[namecnt].items():
                        if href[0] == 'href':
                            disease_address = BKMY_address+href[1]
                            print disease_address

                    bk_address = "http://baike.baidu.com/search/word?word="
                    bk_address += name
                    copy11 = ExtractorWebsite(bk_address)
                    baike_address = copy11.extractor_baike()
                    print baike_address
                    ##---------------------------------------------------##
                    qqyy_address = "http://so.qqyy.com/j?wd="
                    qqyy_address += name
                    
                    ##---------------------------------------------------##
                    row_num+=1
                    ws1.cell(row=row_num,column=0).value = department_name[cnt]
                    ws1.cell(row=row_num,column=1).value = name
                    ws1.cell(row=row_num,column=2).value = disease_address
                    ws1.cell(row=row_num,column=3).value = baike_address
                    ws1.cell(row=row_num,column=4).value = qqyy_address
            #print name_cnt
            ewb1.save(filename=new_filename)
        else:
            print "没有相关疾病"
            #print 0
            row_num+=1
            ws1.cell(row=row_num,column=0).value = department_name[cnt]
            ws1.cell(row=row_num,column=1).value = "没有相关疾病"
            ws1.cell(row=row_num,column=2).value = "None"
            ws1.cell(row=row_num,column=3).value = "None"
            ws1.cell(row=row_num,column=4).value = "None"
            ewb1.save(filename=new_filename)
        

if __name__ == "__main__":
    Start()
