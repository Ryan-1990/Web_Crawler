# -*- coding: utf-8 -*-
#Author: Nan LI @ Tianjin University
#功能描述：抓取好大夫网站的医院地址以及各科室的专家页面地址

import sys
import time
import urllib2
import requests
import lxml.html
import random
import string
import re
import xlrd
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from xlutils.copy import copy


# function: get the directory of script that calls this function
# usage: path = script_path()
def script_path():
    import inspect, os
    caller_file = inspect.stack()[1][1]         # caller's filename
    return os.path.abspath(os.path.dirname(caller_file))# path

def create_folder():
    import os
    path = script_path()
    title = "Hao_Dai_Fu"
    new_path = os.path.join(path, title)
    if not os.path.isdir(new_path):
        os.makedirs(new_path)
        

class Search:
    def __init__(self, address):
        #print address
        self.address = address

        randnum = random.randint(1,9)
        if randnum == 1:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/33.0.1750.154 Safari/537.36'}
        elif randnum == 2:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)',
                           'Referer': 'http://www.baidu.com'}
        elif randnum == 3:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Win64; x64; Trident/4.0)'}
        elif randnum == 4:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)'}
        elif randnum == 5:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0)'}
        elif randnum == 6:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1'}
        elif randnum == 7:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12'}
        elif randnum == 8:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; .NET CLR 2.0.50727; Maxthon 2.0'}
        else:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) AppleWebKit/525.13 (KHTML, like Gecko) Version/3.1 Safari/525.13'}

        #sleepseconds = random.uniform(0, 2)   # Sleep randomly from 0 sec to 5.0 sec. before launching another crawling.
        #time.sleep(sleepseconds)

        req = requests.Request(method='GET',
                               url=self.address,
                               headers=self.header,
                               cookies=None)
        reqprep = req.prepare()
        s = requests.Session()
        resp = s.send(reqprep)

        self.text = resp.text
        self.tree = lxml.html.fromstring(resp.text)
        
class ExtractorWebsite(Search):
    def extractor_website(self):
        a = self.tree.xpath('//div[@class="panelB_blue"]//div[@class="lt"]//table[@id="hosbra"]//td[@width="50%"]/a')
        if len(a) == 0:
            return False
        return a

    def extractor_doctor(self):
        a = self.tree.xpath('//div[@class="box_b"]//div[@class="p_bar"]/a')
        if len(a) == 0:
            return False
        return a    
        



def Start():
    create_folder()
    ScriptPath = script_path()
    HDF_Path = ScriptPath+"/Hao_Dai_Fu/"
    flag = 0

    fname = HDF_Path+"/Hospitals.xlsx"
    new_filename=HDF_Path+'/Doctors.xls'

    #---------打开原文件--------
    bk = xlrd.open_workbook(fname) 

    try:  #已经有文件
        bk1 = xlrd.open_workbook(new_filename, formatting_info=True)
    except: #还没有文件 说明是第一次运行程序
        flag = 1

    sh = bk.sheet_by_name("Sheet1")
    nrows = sh.nrows #获取行数

    if flag == 1:
        #---------新建一个文件--------
        wb1 = Workbook()
        ewb1 = ExcelWriter(workbook=wb1)

        ws1=wb1.worksheets[0]
        ws1.title="Sheet1"
        ws1.cell(row=0,column=0).value = "医院名称"
        ws1.cell(row=0,column=1).value = "科室名称"
        ws1.cell(row=0,column=2).value = "医生链接"

        row_num = 0

        for i in xrange(1,nrows):
            hos_name = sh.cell_value(i,0)
            print hos_name

            hos_address = sh.cell_value(i,1)
            copy11 = ExtractorWebsite(hos_address)
            department_list = copy11.extractor_website()
            count = len(department_list)
            print count

            for cnt in xrange(0,count):
                dep_name = department_list[cnt].text
                print dep_name

                for href in department_list[cnt].items():
                    if href[0] == 'href':
                        dep_address = href[1]
                        #print dep_address

                copy1 = ExtractorWebsite(dep_address)
                doctor_list = copy1.extractor_doctor()
                if doctor_list == False:
                    print 1
                    doctor_address = dep_address.replace('.htm','/menzhen.htm')
                    print doctor_address
                    row_num += 1
                    ws1.cell(row=row_num,column=0).value = hos_name
                    ws1.cell(row=row_num,column=1).value = dep_name
                    ws1.cell(row=row_num,column=2).value = doctor_address

                else:
                    page_cnt = len(doctor_list)-3
                    print page_cnt
                    doctor_address = dep_address.replace('.htm','/menzhen.htm')
                    print doctor_address
                    row_num += 1
                    ws1.cell(row=row_num,column=0).value = hos_name
                    ws1.cell(row=row_num,column=1).value = dep_name
                    ws1.cell(row=row_num,column=2).value = doctor_address

                    for page_num in xrange(2,page_cnt+1):
                        doctor_address = dep_address.replace('.htm','/menzhen_%d.htm'%page_num)
                        print doctor_address
                        row_num += 1
                        ws1.cell(row=row_num,column=0).value = hos_name
                        ws1.cell(row=row_num,column=1).value = dep_name
                        ws1.cell(row=row_num,column=2).value = doctor_address

            ws1.cell(row=0,column=100).value = i
            ws1.cell(row=1,column=100).value = row_num
            ewb1.save(filename=new_filename)

    else:
        sh1 = bk1.sheet_by_name("Sheet1")
        wb = copy(bk1)
        ws = wb.get_sheet(0)

        start = int(sh1.cell_value(0,100))
        row_num = int(sh1.cell_value(1,100))

        for i in xrange(start+1,nrows):
            hos_name = sh.cell_value(i,0)
            print hos_name

            hos_address = sh.cell_value(i,1)
            copy11 = ExtractorWebsite(hos_address)
            department_list = copy11.extractor_website()
            count = len(department_list)
            print count

            for cnt in xrange(0,count):
                dep_name = department_list[cnt].text
                print dep_name

                for href in department_list[cnt].items():
                    if href[0] == 'href':
                        dep_address = href[1]
                        #print dep_address

                copy1 = ExtractorWebsite(dep_address)
                doctor_list = copy1.extractor_doctor()
                if doctor_list == False:
                    print 1
                    doctor_address = dep_address.replace('.htm','/menzhen.htm')
                    print doctor_address
                    row_num += 1
                    ws.write(row_num, 0, hos_name)
                    ws.write(row_num, 1, dep_name)
                    ws.write(row_num, 2, doctor_address)

                else:
                    page_cnt = len(doctor_list)-3
                    print page_cnt
                    doctor_address = dep_address.replace('.htm','/menzhen.htm')
                    print doctor_address
                    row_num += 1
                    ws.write(row_num, 0, hos_name)
                    ws.write(row_num, 1, dep_name)
                    ws.write(row_num, 2, doctor_address)

                    for page_num in xrange(2,page_cnt+1):
                        doctor_address = dep_address.replace('.htm','/menzhen_%d.htm'%page_num)
                        print doctor_address
                        row_num += 1
                        ws.write(row_num, 0, hos_name)
                        ws.write(row_num, 1, dep_name)
                        ws.write(row_num, 2, doctor_address)

            ws.write(0, 100, i)
            ws.write(1, 100, row_num)
            wb.save(new_filename)

       

if __name__ == "__main__":
    Start()
