# -*- coding: utf-8 -*-
#Author: Nan LI @ Tianjin University
#功能描述：1.从给定excel表格取出第一列作为百度搜索关键字，抓取网页链接地址，并写入新表格中
#         2.从表格中提取药品名称，放入指定的几个药品网站搜索，抓取器搜索结果页面内容
import sys
import time
import xlrd
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
import urllib2
import requests
import lxml.html
import random

###############################################################################################################
Save_Freq = 100 #定义保存数据的频率，每Save_Freq个数据保存一次文件
###############################################################################################################

# function: get the directory of script that calls this function
# usage: path = script_path()
def script_path():
    import inspect, os
    caller_file = inspect.stack()[1][1]         # caller's filename
    return os.path.abspath(os.path.dirname(caller_file))# path

def create_folder():
    import os
    path = script_path()
    title = ["Bai_Ke", "Quan_Min_Jian_Kang", "Xun_Yi_Wen_Yao", "Yao_Pin_Tong"]
    for num in range(0,len(title)):
        new_path = os.path.join(path, title[num])
        if not os.path.isdir(new_path):
            os.makedirs(new_path)

class Search:
    def __init__(self, address, keyword, encoding):
        #print address
        self.address = address
        self.keyword = keyword
        self.encoding = encoding

        randnum = random.randint(1,9)
        if randnum == 1:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/33.0.1750.154 Safari/537.36'}
        elif randnum == 2:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)',
                           'Referer': 'http://www.baidu.com'}
        elif randnum == 3:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Win64; x64; Trident/4.0)'}
        elif randnum == 4:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)'}
        elif randnum == 5:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0)'}
        elif randnum == 6:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1'}
        elif randnum == 7:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12'}
        elif randnum == 8:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; .NET CLR 2.0.50727; Maxthon 2.0'}
        else:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) AppleWebKit/525.13 (KHTML, like Gecko) Version/3.1 Safari/525.13'}

        sleepseconds = random.uniform(0, 1)   # Sleep randomly from 0 sec to 1.0 sec. before launching another crawling.
        time.sleep(sleepseconds)

        #data = requests.get(self.address, headers=self.header).text   # This method is easy to be blocked.

        req = requests.Request(method='GET',
                               url=self.address,
                               params=self.keyword,
                               headers=self.header,
                               cookies=None)
        reqprep = req.prepare()
        s = requests.Session()
        resp = s.send(reqprep)
        if self.encoding:
            resp.encoding = 'utf-8'
            self.encoding = resp.encoding
        self.text = resp.text
        self.tree = lxml.html.fromstring(resp.text)

class ExtractorWebsite(Search):

    def get_url(self, items):
        for i in items:
           if i[0] == 'href':
              return i[1]

    def extractor_website(self):
        a = self.tree.xpath('//div[@class="c-gap-top"]/a')
        if len(a) == 0:
            return False
        return self.get_url(a[0].items())
    
    def extractor_baike(self):
        b = self.tree.xpath('//div[@class="mod-list"]//a')
        if len(b) == 0:
            return False
        return self.get_url(b[0].items())

def Start():
    #reload(sys)
    #sys.setdefaultencoding('utf-8')
    
    create_folder()
    ScriptPath = script_path()
    BK_Path = ScriptPath+"/Bai_Ke/"
    QMJK_Path = ScriptPath+"/Quan_Min_Jian_Kang/"
    XYWY_Path = ScriptPath+"/Xun_Yi_Wen_Yao/"
    YPT_Path = ScriptPath+"/Yao_Pin_Tong/"

    baidu_address = "http://www.baidu.com/s"
######################################################################
    #---------打开原文件--------
    fname = ScriptPath+"/Drugs_Database_WD_20140323_v1.xlsx"
    bk = xlrd.open_workbook(fname)
    try:
        sh = bk.sheet_by_name("Sheet1")
    except:
        print "no sheet in %s named Sheet1" % fname
    nrows = sh.nrows #获取行数
    ncols = sh.ncols #获取列数

######################################################################
    #---------新建一个文件--------
    wb1 = Workbook()
    ewb1 = ExcelWriter(workbook=wb1)
    new_filename=ScriptPath+'/Drugs_Database_WD_20140323_v1_new.xlsx'
    ws1=wb1.worksheets[0]
    ws1.title="Sheet1"

######################################################################
    address = "http://www.baidu.com/s"
    for col in range(0,ncols):
        ws1.cell(row=0,column=col).value = sh.cell_value(0,col)
    ws1.cell(row=0,column=8).value = "百度药品链接"
    for i in xrange(1,nrows):
    #for i in xrange(23101,nrows):

        for j in range(0,ncols):
            ws1.cell(row=i,column=j).value = sh.cell_value(i,j)

        key = sh.cell_value(i,0).encode('utf8')
        print key
        keyword = {'wd': key, 'rn': '1'}
        ex = ExtractorWebsite(baidu_address,keyword,0)
        website = ex.extractor_website()
        if website:
            print website
            ws1.cell(row=i,column=8).value = website
        else:
            print "Can't find"
            ws1.cell(row=i,column=8).value = "Can't find"
        if i%Save_Freq==0:
            ewb1.save(filename=new_filename)

        ##############################################
        drug_name = sh.cell_value(i,1)
        drug_name_gbk = drug_name.encode('gbk')
        drug_name_url = urllib2.quote(drug_name_gbk)
        bk_address = "http://baike.baidu.com/search/word?word="
        bk_address += drug_name
        copy1 = ExtractorWebsite(bk_address,'',0)
        new_address = copy1.extractor_baike()
        print new_address
        copy11 = ExtractorWebsite(new_address,'',1)
        spath = BK_Path+"BK_"+'%d'%(i+1)+".html"
        f=open(spath,"w")
        f.write(copy11.text.encode('utf8'))
        f.close()
            
        qm_120_address = "http://yp.qm120.com/search.aspx?keys="
        qm_120_address += drug_name_url
        #print qm_120_address
        copy2 = ExtractorWebsite(qm_120_address,'',0)
        spath = QMJK_Path+"QMJK_"+'%d'%(i+1)+".html"
        f=open(spath,"w")
        f.write(copy2.text.encode('utf8'))
        f.close()

        xywy_address = "http://yao.xywy.com/so/?q="
        xywy_address += drug_name
        #print xywy_address
        copy3 = ExtractorWebsite(xywy_address,'',0)
        spath = XYWY_Path+"XYWY_"+'%d'%(i+1)+".html"
        f=open(spath,"w")
        f.write(copy3.text.encode('utf8'))
        f.close()

        ypt_address = "http://ypk.39.net/search/all?k="
        ypt_address += drug_name_url
        #print ypt_address
        copy4 = ExtractorWebsite(ypt_address,'',0)
        spath = YPT_Path+"YPT_"+'%d'%(i+1)+".html"
        f=open(spath,"w")
        f.write(copy4.text.encode('utf8'))
        f.close()

    ewb1.save(filename=new_filename)


if __name__ == "__main__":
    Start()
