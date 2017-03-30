# -*- coding: utf-8 -*-
#Author: Nan LI @ Tianjin University
#功能描述：抓取好大夫网站的医院地址以及各科室的专家页面地址

import sys
import time
import urllib2
import requests
import lxml.html
import random
import string
import re
import xlrd
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter


# function: get the directory of script that calls this function
# usage: path = script_path()
def script_path():
    import inspect, os
    caller_file = inspect.stack()[1][1]         # caller's filename
    return os.path.abspath(os.path.dirname(caller_file))# path

def create_folder():
    import os
    path = script_path()
    title = "Hao_Dai_Fu"
    new_path = os.path.join(path, title)
    if not os.path.isdir(new_path):
        os.makedirs(new_path)
        

class Search:
    def __init__(self, address):
        #print address
        self.address = address

        randnum = random.randint(1,9)
        if randnum == 1:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/33.0.1750.154 Safari/537.36'}
        elif randnum == 2:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)',
                           'Referer': 'http://www.baidu.com'}
        elif randnum == 3:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Win64; x64; Trident/4.0)'}
        elif randnum == 4:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; Trident/4.0)'}
        elif randnum == 5:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.1; WOW64; Trident/4.0)'}
        elif randnum == 6:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) Gecko/2008070208 Firefox/3.0.1'}
        elif randnum == 7:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.1) Gecko/20070803 Firefox/1.5.0.12'}
        elif randnum == 8:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 5.1; .NET CLR 2.0.50727; Maxthon 2.0'}
        else:
            self.header = {'Accept-Charset': 'utf-8', 'Accept': 'text/css,*/*;q=0.1',
                           'User-Agent': 'Mozilla/5.0 (Windows; U; Windows NT 5.2) AppleWebKit/525.13 (KHTML, like Gecko) Version/3.1 Safari/525.13'}

        #sleepseconds = random.uniform(0, 2)   # Sleep randomly from 0 sec to 5.0 sec. before launching another crawling.
        #time.sleep(sleepseconds)

        req = requests.Request(method='GET',
                               url=self.address,
                               headers=self.header,
                               cookies=None)
        reqprep = req.prepare()
        s = requests.Session()
        resp = s.send(reqprep)

        self.text = resp.text
        self.tree = lxml.html.fromstring(resp.text)
        
class ExtractorWebsite(Search):
    def extractor_website(self):
        a = self.tree.xpath('//table[@class="tab_white"]//span[@class="font14"]')
        if len(a) == 0:
            return False
        return a

    def extractor_name(self):
        a = self.tree.xpath('//table[@class="list"]//span[@class="font16"]/a')
        if len(a) == 0:
            return False
        return a    
        



def Start():
    create_folder()
    ScriptPath = script_path()
    XYWY_Path = ScriptPath+"/Xun_Yi_Wen_Yao/"
    HDF_Path = ScriptPath+"/Hao_Dai_Fu/"

    #---------打开原文件--------

    fname = XYWY_Path+"/Hospital_Names.xlsx"

    bk = xlrd.open_workbook(fname)

    try:

        sh = bk.sheet_by_name("Sheet1")

    except:

        print "no sheet in %s named Sheet1" % fname

    nrows = sh.nrows #获取行数

    
    #---------新建一个文件--------
    wb1 = Workbook()
    ewb1 = ExcelWriter(workbook=wb1)
    new_filename=HDF_Path+'/Hospitals.xlsx'
    ws1=wb1.worksheets[0]
    ws1.title="Sheet1"
    ws1.cell(row=0,column=0).value = "医院名称"
    ws1.cell(row=0,column=1).value = "医院链接"
 
    HDF_address = "http://so.haodf.com/index/search?type=hospital&kw="

    row_num = 0
    
    for i in xrange(1,nrows):

        key = sh.cell_value(i,2)

        print key

 
        copy = ExtractorWebsite(HDF_address+key.encode('gbk'))

        #print copy.text

        if copy.extractor_website() == 0:
            temp = copy.extractor_name()
            hospital_name = temp[0].text.encode('utf8')
            row_num += 1
            ws1.cell(row=row_num,column=0).value = hospital_name

            print hospital_name
            
            for href in temp[0].items():
                if href[0] == 'href':
                    hospital_address = href[1]
            print hospital_address + '\n'
            ws1.cell(row=row_num,column=1).value = hospital_address

            if row_num%100==0:
                ewb1.save(filename=new_filename)

        else:
            print "Can't find\n"

    ewb1.save(filename=new_filename)
       

if __name__ == "__main__":
    Start()
